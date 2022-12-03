import os
import sys
from pathlib import Path

import openpyxl
from PyQt5 import QtGui
from PyQt5.QtWidgets import *
from PyQt5.uic import loadUi
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


class Main(QDialog):
    def __init__(self):
        super(Main, self).__init__()
        loadUi('main.ui', self)
        self.setWindowTitle('Работа с файлами в Python')

        self.btn_wrd.clicked.connect(self.execute_word)
        self.btn_xl.clicked.connect(self.execute_xlsx)

    def execute_word(self):
        document_path = Path(__file__).parent / "example.docx"
        doc = DocxTemplate(document_path)
        context = {"FIO": self.lineEdit_6.text(),
                   "AdresComp": self.lineEdit_2.text(),
                   "Adres": self.lineEdit_3.text(),
                   "email": self.lineEdit_4.text(),
                   "Web": self.lineEdit_5.text()}
        doc.render(context)
        doc.save(Path(__file__).parent / "result_example.docx")
        os.system('start result_example.docx')

    def execute_xlsx(self):
        fn = 'example.xlsx'
        wb = load_workbook(fn)
        ws = wb['data']
        ws['A1'] = "            " + self.lineEdit_6.text()
        ws['D6'] = self.lineEdit_2.text()
        ws['D7'] = self.lineEdit_3.text()
        ws['D8'] = self.lineEdit_4.text()
        ws['D9'] = self.lineEdit_5.text()

        wb.save(Path(__file__).parent / "result_example.xlsx")
        wb.close()
        os.system('start result_example.xlsx')


def main():
    app = QApplication(sys.argv)
    window = Main()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
